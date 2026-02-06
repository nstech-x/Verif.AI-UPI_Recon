import os
import csv
import json
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Dict, Optional
from config import OUTPUT_DIR

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _ensure_run_dirs(run_id: str):
    base = os.path.join(OUTPUT_DIR, run_id)
    # Main directories
    for sub in ('reports', 'ttum', 'annexure', 'audit'):
        os.makedirs(os.path.join(base, sub), exist_ok=True)

    # Segmented subdirectories under reports/
    reports_base = os.path.join(base, 'reports')
    for sub in ('ttum & annex', 'listing', 'reconciliation', 'legacy'):
        os.makedirs(os.path.join(reports_base, sub), exist_ok=True)


# Removed _format_value to preserve native numeric/date types in outputs


def write_report(run_id: str, cycle_id: Optional[str], subdir: str, filename: str, headers: List[str], rows: List[Dict]):
    """Write a CSV report under OUTPUT_DIR/<run_id>/<subdir>/[cycle_<cycle_id>/]filename.

    - headers: ordered list of column names
    - rows: list of dicts; values will be formatted according to key heuristics
    Ensures UTF-8 (no BOM) and newline-safe writing on Windows.
    """
    _ensure_run_dirs(run_id)
    base = os.path.join(OUTPUT_DIR, run_id, subdir)
    if cycle_id:
        base = os.path.join(base, f"cycle_{cycle_id}")
    os.makedirs(base, exist_ok=True)

    out_path = os.path.join(base, filename)

    # Write CSV with exact header order and UTF-8 encoding
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in rows:
            row = []
            for h in headers:
                row.append(r.get(h))
            writer.writerow(row)
        # Explicitly flush to ensure all data is written
        f.flush()
        os.fsync(f.fileno())

    return out_path


def write_ttum_xlsx(run_id: str, cycle_id: Optional[str], filename: str, headers: List[str], rows: List[Dict]) -> str:
    """Write TTUM data to XLSX file using openpyxl.
    
    Args:
        run_id: Run identifier
        cycle_id: Optional cycle identifier
        filename: Output filename (without extension)
        headers: List of column headers
        rows: List of dictionaries with row data
    
    Returns:
        Path to created XLSX file
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")
    
    _ensure_run_dirs(run_id)
    base = os.path.join(OUTPUT_DIR, run_id, 'ttum')
    if cycle_id:
        base = os.path.join(base, f"cycle_{cycle_id}")
    os.makedirs(base, exist_ok=True)
    
    out_path = os.path.join(base, f"{filename}.xlsx")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "TTUM"
    
    # Write headers with formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write data rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header)
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save and close workbook to flush buffers
    wb.save(out_path)
    try:
        wb.close()
    except Exception:
        pass
    return out_path


def write_ttum_csv(run_id: str, cycle_id: Optional[str], filename: str, headers: List[str], rows: List[Dict]) -> str:
    """Write TTUM data to CSV file.
    
    Args:
        run_id: Run identifier
        cycle_id: Optional cycle identifier
        filename: Output filename (without extension)
        headers: List of column headers
        rows: List of dictionaries with row data
    
    Returns:
        Path to created CSV file
    """
    _ensure_run_dirs(run_id)
    base = os.path.join(OUTPUT_DIR, run_id, 'ttum')
    if cycle_id:
        base = os.path.join(base, f"cycle_{cycle_id}")
    os.makedirs(base, exist_ok=True)
    
    out_path = os.path.join(base, f"{filename}.csv")

    # Write CSV with exact header order and UTF-8 encoding
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row_data in rows:
            row = []
            for header in headers:
                row.append(row_data.get(header))
            writer.writerow(row)
        # Explicitly flush to ensure all data is written
        f.flush()
        os.fsync(f.fileno())
    
    return out_path


def write_ttum_pandas(run_id: str, cycle_id: Optional[str], filename: str, headers: List[str], rows: List[Dict], format: str = 'xlsx') -> str:
    """Write TTUM data using pandas (faster for large datasets).
    
    Args:
        run_id: Run identifier
        cycle_id: Optional cycle identifier
        filename: Output filename (without extension)
        headers: List of column headers
        rows: List of dictionaries with row data
        format: Output format ('xlsx' or 'csv')
    
    Returns:
        Path to created file
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("pandas is required for this function. Install with: pip install pandas")
    
    _ensure_run_dirs(run_id)
    base = os.path.join(OUTPUT_DIR, run_id, 'ttum')
    if cycle_id:
        base = os.path.join(base, f"cycle_{cycle_id}")
    os.makedirs(base, exist_ok=True)
    
    # Pass raw rows directly; let pandas preserve dtypes
    df = pd.DataFrame(rows, columns=headers)
    
    if format.lower() == 'xlsx':
        out_path = os.path.join(base, f"{filename}.xlsx")
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='TTUM', index=False)
            # Auto-adjust column widths
            worksheet = writer.sheets['TTUM']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    else:  # csv
        out_path = os.path.join(base, f"{filename}.csv")
        df.to_csv(out_path, index=False, encoding='utf-8-sig')
    
    return out_path


def get_ttum_files(run_id: str, cycle_id: Optional[str] = None, format: str = 'all') -> List[str]:
    """Get list of TTUM files for a run.
    
    Args:
        run_id: Run identifier
        cycle_id: Optional cycle identifier to filter
        format: File format to return ('csv', 'xlsx', 'json', or 'all')
    
    Returns:
        List of file paths
    """
    files = set()
    for subdir in ['ttum', 'reports']:
        base_dir = os.path.join(OUTPUT_DIR, run_id, subdir)
        if cycle_id:
            base_dir = os.path.join(base_dir, f"cycle_{cycle_id}")
        
        if not os.path.exists(base_dir):
            continue
        
        for filename in os.listdir(base_dir):
            if any(keyword in filename for keyword in ['ttum', 'unmatched', 'exceptions']):
                filepath = os.path.join(base_dir, filename)
                if os.path.isfile(filepath):
                    if format == 'all':
                        files.add(filepath)
                    elif format == 'csv' and filename.endswith('.csv'):
                        files.add(filepath)
                    elif format == 'xlsx' and filename.endswith('.xlsx'):
                        files.add(filepath)
                    elif format == 'json' and filename.endswith('.json'):
                        files.add(filepath)
    
    return sorted(list(files))
